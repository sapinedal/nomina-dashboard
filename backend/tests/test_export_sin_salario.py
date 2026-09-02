"""Listado de empleados activos sin salario en Trazalo.

Dos cosas que probar sin tocar Trazalo: que el libro sale bien armado y que el
filtro por áreas es fail-closed. La consulta a Trazalo en sí no se prueba aquí
(necesita esa base); lo que se prueba es todo lo que pasa DESPUÉS de traerla.
"""
import unittest


def _emp(cedula, nombre, area, cargo="AUXILIAR", sede="PRINCIPAL"):
    return {"cedula": cedula, "nombre": nombre, "area": area,
            "cargo": cargo, "sede": sede}


class TestFiltroPorAreas(unittest.TestCase):
    """El listado lleva identidad de personal: no puede ser más laxo que el
    resto del dashboard."""

    @classmethod
    def setUpClass(cls):
        try:
            from app.services.dashboard_service import filtrar_filas_por_areas
        except ImportError as exc:
            raise unittest.SkipTest(f"deps no disponibles en este entorno ({exc})")
        cls.filtrar = staticmethod(filtrar_filas_por_areas)
        cls.filas = [_emp("1", "ANA", "NOMINA"), _emp("2", "LUIS", "SST"),
                     _emp("3", "PEPE", "CIRUGIA")]

    def test_admin_ve_todo(self):
        """None es el centinela de admin: sin restricción."""
        self.assertEqual(len(self.filtrar(self.filas, None)), 3)

    def test_restringido_solo_ve_sus_areas(self):
        r = self.filtrar(self.filas, ["NOMINA", "SST"])
        self.assertEqual([f["cedula"] for f in r], ["1", "2"])

    def test_restringido_sin_areas_no_ve_nada(self):
        """El caso que ya se equivocó una vez: lista vacía es fail-closed, no
        'sin filtro'. Abrirlo es una fuga de datos entre áreas."""
        self.assertEqual(self.filtrar(self.filas, []), [])

    def test_un_area_ajena_no_se_cuela(self):
        self.assertEqual(self.filtrar(self.filas, ["COMPRAS"]), [])


class TestLibroSinSalario(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
            import xlsxwriter  # noqa: F401
            from app.services.excel_report import construir_libro_sin_salario
        except ImportError as exc:
            raise unittest.SkipTest(f"deps no disponibles en este entorno ({exc})")
        cls.construir = staticmethod(construir_libro_sin_salario)

    def _hoja(self, filas):
        import openpyxl
        return openpyxl.load_workbook(self.construir(filas)).active

    def test_lleva_el_aviso_de_la_consecuencia(self):
        """Sin decir qué implica el hueco, el listado parece una curiosidad y no
        una tarea pendiente."""
        ws = self._hoja([_emp("1", "ANA", "NOMINA")])
        aviso = ws.cell(1, 1).value or ""
        self.assertIn("ACTIVOS", aviso)
        self.assertIn("excluidos", aviso)

    def test_columnas_y_datos(self):
        ws = self._hoja([_emp("1", "ANA GOMEZ", "NOMINA", "AUXILIAR", "SEDE 1")])
        self.assertEqual([ws.cell(4, c).value for c in range(1, 6)],
                         ["Cédula", "Nombre Empleado", "Área", "Cargo", "Sede"])
        self.assertEqual([ws.cell(5, c).value for c in range(1, 6)],
                         ["1", "ANA GOMEZ", "NOMINA", "AUXILIAR", "SEDE 1"])

    def test_la_cedula_va_como_texto(self):
        """Como número pierde los ceros a la izquierda y deja de cruzar con los
        demás sistemas."""
        ws = self._hoja([_emp("0012345", "ANA", "NOMINA")])
        self.assertEqual(ws.cell(5, 1).value, "0012345")

    def test_el_total_cuenta_las_filas(self):
        ws = self._hoja([_emp(str(i), f"E{i}", "NOMINA") for i in range(7)])
        self.assertEqual(ws.cell(ws.max_row, 1).value, "TOTAL")
        self.assertEqual(ws.cell(ws.max_row, 2).value, 7)

    def test_sin_nadie_sigue_siendo_un_libro_valido(self):
        """Que no falte nadie es un resultado legítimo, no un error."""
        ws = self._hoja([])
        self.assertEqual(ws.cell(ws.max_row, 2).value, 0)

    def test_un_campo_nulo_no_revienta(self):
        """Trazalo tiene ~570 empleados sin área asignada: no pueden tumbar el
        reporte que existe justamente para señalar huecos de datos."""
        ws = self._hoja([{"cedula": "1", "nombre": None, "area": None,
                          "cargo": None, "sede": None}])
        self.assertEqual(ws.cell(5, 2).value, "")


if __name__ == "__main__":
    unittest.main()
