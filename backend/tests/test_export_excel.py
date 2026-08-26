"""El Excel no debe mezclar horas con dias en una sola columna.

La columna `dias` de la tabla es una cantidad cuya unidad depende de `unidad`:
son HORAS en recargos y horas extras, y DIAS en incapacidades, vacaciones y
licencias. El export las volcaba todas bajo un unico encabezado "Dias", de modo
que 0.33 (veinte minutos de hora extra) y 22 (dias de vacaciones) aparecian en
la misma columna como si fueran sumables.

Se omite si faltan las dependencias, igual que el resto de la suite.
"""
import unittest
from datetime import date


def _fila(**kw):
    base = dict(
        cedula="1", nombre_empleado="ANA", area="CONTACT CENTER", cargo="AUX",
        tipo_novedad="", categoria="", fecha_inicio=date(2026, 8, 25),
        fecha_fin=date(2026, 8, 25), dias=0.0, unidad="dias",
        valor_calculado=0.0, periodo="2026-08", estado=None,
        archivo_origen="c.xlsx", hoja_origen="h1",
    )
    base.update(kw)
    return base


class TestLibroExcelSegregado(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
            import xlsxwriter  # noqa: F401
            from app.services.excel_report import construir_libro_excel
        except ImportError as exc:
            raise unittest.SkipTest(f"deps no disponibles en este entorno ({exc})")
        cls.construir = staticmethod(construir_libro_excel)
        cls.filas = [
            _fila(tipo_novedad="HORAS EXTRAS DIURNAS", categoria="H. Extras & Recargos",
                  dias=0.33, unidad="horas", valor_calculado=5156.25),
            _fila(tipo_novedad="RECARGO NOCTURNO", categoria="H. Extras & Recargos",
                  dias=8.0, unidad="horas", valor_calculado=35000.0),
            _fila(tipo_novedad="VACACIONES", categoria="Vacaciones",
                  dias=22.0, unidad="dias", valor_calculado=2200000.0),
            _fila(tipo_novedad="INCAPACIDAD", categoria="Incapacidades",
                  dias=2.0, unidad="dias", valor_calculado=200000.0),
        ]

    def _abrir(self):
        import openpyxl
        return openpyxl.load_workbook(self.construir(self.filas))

    def test_tres_hojas_segregadas(self):
        self.assertEqual(
            self._abrir().sheetnames,
            ["Novedades", "Horas extras y recargos", "Ausencias y días"],
        )

    def test_horas_y_dias_nunca_en_la_misma_celda(self):
        """El nucleo del defecto: una fila aporta a Horas o a Dias, jamas a ambas."""
        ws = self._abrir()["Novedades"]
        cab = [c.value for c in ws[1]]
        iH, iD = cab.index("Horas") + 1, cab.index("Días") + 1
        for r in range(2, ws.max_row):  # max_row es la fila TOTAL
            horas, dias = ws.cell(r, iH).value, ws.cell(r, iD).value
            with self.subTest(fila=r):
                self.assertFalse(horas is not None and dias is not None,
                                 "una novedad no puede tener horas y dias a la vez")
                self.assertFalse(horas is None and dias is None,
                                 "toda novedad debe aportar su cantidad a alguna columna")

    def test_totales_por_unidad(self):
        ws = self._abrir()["Novedades"]
        cab = [c.value for c in ws[1]]
        tot = ws.max_row
        self.assertEqual(ws.cell(tot, 1).value, "TOTAL")
        self.assertAlmostEqual(ws.cell(tot, cab.index("Horas") + 1).value, 8.33, places=2)
        self.assertAlmostEqual(ws.cell(tot, cab.index("Días") + 1).value, 24.0, places=2)
        self.assertAlmostEqual(ws.cell(tot, cab.index("Valor (COP)") + 1).value,
                               2440156.25, places=2)

    def test_cada_hoja_solo_trae_su_unidad(self):
        wb = self._abrir()
        horas = [c.value for c in wb["Horas extras y recargos"][1]]
        dias = [c.value for c in wb["Ausencias y días"][1]]
        self.assertIn("Horas", horas)
        self.assertNotIn("Días", horas)
        self.assertIn("Días", dias)
        self.assertNotIn("Horas", dias)

    def test_export_de_un_panel_trae_una_sola_hoja(self):
        """Desde Ausentismo u Horas extras el libro no debe traer las tres
        hojas: las otras dos saldrian vacias o duplicando lo mismo."""
        import openpyxl
        solo_horas = [f for f in self.filas if f["unidad"] == "horas"]
        solo_dias = [f for f in self.filas if f["unidad"] == "dias"]

        wb = openpyxl.load_workbook(self.construir(solo_horas, panel="horas-extras"))
        self.assertEqual(wb.sheetnames, ["Horas extras y recargos"])
        self.assertNotIn("Días", [c.value for c in wb.active[1]])

        wb = openpyxl.load_workbook(self.construir(solo_dias, panel="ausentismo"))
        self.assertEqual(wb.sheetnames, ["Ausencias y días"])
        self.assertNotIn("Horas", [c.value for c in wb.active[1]])

    def test_sin_panel_sigue_trayendo_las_tres_hojas(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.construir(self.filas, panel=None))
        self.assertEqual(len(wb.sheetnames), 3)

    def test_el_valor_llega_a_la_hoja(self):
        """`valor` se guarda nulo en la carga; el importe lo calcula la consulta.
        Si el export volviera a leer la columna cruda, esto se caeria."""
        ws = self._abrir()["Novedades"]
        cab = [c.value for c in ws[1]]
        iV = cab.index("Valor (COP)") + 1
        valores = [ws.cell(r, iV).value for r in range(2, ws.max_row)]
        self.assertTrue(all(v is not None and v > 0 for v in valores), valores)


if __name__ == "__main__":
    unittest.main()
