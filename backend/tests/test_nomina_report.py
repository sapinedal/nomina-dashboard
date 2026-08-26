"""Liquidación de incapacidades por tramos.

Son cifras de dinero con efecto legal, así que cada tramo se comprueba con un
número calculado a mano, no con lo que devuelva la función.

Sobre un salario de 3.000.000 el día vale 100.000, y sobre un SMLMV de
1.300.000 el piso diario es 43.333,33.
"""
import unittest

from app.services.nomina_report import liquidar_incapacidad

SALARIO = 3_000_000.0     # día = 100.000
SMLMV = 1_300_000.0       # piso diario = 43.333,33
DIA_66 = 0.6667 * (SALARIO / 30)   # 66.670
DIA_50 = 0.50 * (SALARIO / 30)     # 50.000


class TestLiquidarIncapacidad(unittest.TestCase):

    def test_primeros_dos_dias_al_66(self):
        """Días 1-2 los paga el empleador, pero al mismo porcentaje."""
        r = liquidar_incapacidad(0, 2, SALARIO, SMLMV)
        self.assertAlmostEqual(r["valor"], 2 * DIA_66, places=2)
        self.assertEqual(r["dias_66"], 2)
        self.assertEqual(r["dias_50"], 0)

    def test_novedad_que_cruza_el_dia_90_se_parte(self):
        """El caso que se equivoca si se liquida la novedad de una sola vez:
        quien va por el día 89 y pide 5 días cobra 1 al 66,67% y 4 al 50%."""
        r = liquidar_incapacidad(89, 5, SALARIO, SMLMV)
        self.assertEqual(r["dias_66"], 1)
        self.assertEqual(r["dias_50"], 4)
        self.assertAlmostEqual(r["valor"], 1 * DIA_66 + 4 * DIA_50, places=2)
        self.assertTrue(any("cruza el dia 90" in o for o in r["observaciones"]))

    def test_tramo_del_50_completo(self):
        r = liquidar_incapacidad(100, 10, SALARIO, SMLMV)
        self.assertEqual(r["dias_66"], 0)
        self.assertEqual(r["dias_50"], 10)
        self.assertAlmostEqual(r["valor"], 10 * DIA_50, places=2)

    def test_piso_de_un_smlmv(self):
        """Con salario mínimo, el 66,67% queda por debajo del piso legal:
        debe pagarse el piso, no el porcentaje."""
        r = liquidar_incapacidad(0, 3, SMLMV, SMLMV)
        self.assertAlmostEqual(r["valor"], 3 * (SMLMV / 30), places=2)
        self.assertGreater(r["valor"], 3 * 0.6667 * (SMLMV / 30))
        self.assertIn("se aplico el piso de 1 SMLMV", r["observaciones"])

    def test_el_piso_tambien_cubre_a_quien_gana_mas_del_minimo(self):
        """La norma no limita el piso a quien devenga el mínimo: ninguna
        incapacidad puede pagarse bajo el salario mínimo diario (Decreto 780 de
        2016 art. 3.2.1.10). Con el mínimo de 2026 el umbral es 2.626.357, así
        que un salario de 2.000.000 cobra el piso y no el 66,67% de su día."""
        salario, smlmv_2026 = 2_000_000.0, 1_750_905.0
        r = liquidar_incapacidad(0, 1, salario, smlmv_2026)
        self.assertAlmostEqual(r["valor"], smlmv_2026 / 30, places=2)
        self.assertGreater(r["valor"], 0.6667 * salario / 30)
        self.assertIn("se aplico el piso de 1 SMLMV", r["observaciones"])

    def test_por_encima_del_umbral_se_liquida_el_porcentaje_de_SU_salario(self):
        """El piso no reemplaza el salario: quien gana por encima del umbral
        cobra el 66,67% de SU día, que es más que el mínimo diario."""
        salario, smlmv_2026 = 3_000_000.0, 1_750_905.0
        r = liquidar_incapacidad(0, 1, salario, smlmv_2026)
        self.assertAlmostEqual(r["valor"], 0.6667 * salario / 30, places=2)
        self.assertEqual(r["observaciones"], [])

    def test_mas_alla_del_dia_180_no_lo_paga_la_empresa(self):
        """Corresponde a la AFP. Se liquida en 0 y se deja constancia."""
        r = liquidar_incapacidad(178, 5, SALARIO, SMLMV)
        self.assertEqual(r["dias_50"], 2)
        self.assertEqual(r["dias_sin_pago"], 3)
        self.assertAlmostEqual(r["valor"], 2 * DIA_50, places=2)
        self.assertTrue(any("AFP" in o for o in r["observaciones"]))

    def test_totalmente_fuera_de_rango(self):
        r = liquidar_incapacidad(200, 10, SALARIO, SMLMV)
        self.assertEqual(r["valor"], 0.0)
        self.assertEqual(r["dias_sin_pago"], 10)

    def test_sin_salario_no_inventa_una_cifra(self):
        """Si el empleado no está en Trazalo no se puede liquidar: se avisa en
        vez de devolver 0 en silencio."""
        r = liquidar_incapacidad(0, 5, None, SMLMV)
        self.assertEqual(r["valor"], 0.0)
        self.assertEqual(r["dias_sin_pago"], 5)
        self.assertTrue(any("sin salario" in o for o in r["observaciones"]))

    def test_cero_dias(self):
        self.assertEqual(liquidar_incapacidad(0, 0, SALARIO, SMLMV)["valor"], 0.0)

    def test_dias_fraccionados(self):
        r = liquidar_incapacidad(0, 0.5, SALARIO, SMLMV)
        self.assertAlmostEqual(r["valor"], 0.5 * DIA_66, places=2)


if __name__ == "__main__":
    unittest.main()


class TestArmarReporte(unittest.TestCase):
    """El neto por empleado: devengado + incapacidad + extras + otros pagos."""

    def _fila(self, **kw):
        base = dict(cedula="1", nombre_empleado="ANA", area="NOMINA", cargo="AUX",
                    salario=SALARIO, valor_extras=0.0, dias_incapacidad=0.0,
                    dias_no_remunerados=0.0, valor_otros_pagos=0.0)
        base.update(kw)
        return base

    def _incap(self, ini, fin, dias, en_periodo=True):
        return dict(fecha_inicio=ini, fecha_fin=fin, dias=dias, en_periodo=en_periodo)

    def test_sin_novedades_cobra_su_salario_completo(self):
        """Un empleado que no tuvo ninguna novedad en el mes debe aparecer en el
        reporte con su salario integro y diferencia cero. Antes ni siquiera
        salia: la consulta arrancaba desde novedades_nomina."""
        from app.services.nomina_report import armar_reporte
        r = armar_reporte([self._fila(num_novedades=0)], {}, SMLMV)[0]
        self.assertEqual(r["dias_efectivos"], 30)
        self.assertAlmostEqual(r["salario_devengado"], SALARIO, places=2)
        self.assertAlmostEqual(r["total_a_pagar"], SALARIO, places=2)
        self.assertAlmostEqual(r["diferencia_vs_salario"], 0.0, places=2)

    def test_javier_con_salario_minimo_no_pierde_nada(self):
        """Caso real: salario = SMLMV, 3 dias de incapacidad. El piso legal
        equivale a su salario diario, asi que cobra el 100%."""
        from app.services.nomina_report import armar_reporte
        minimo = 1_790_000.0
        r = armar_reporte(
            [self._fila(salario=minimo, dias_incapacidad=3.0)],
            {"1": [self._incap("2026-06-23", "2026-06-25", 3.0)]}, minimo)[0]
        self.assertAlmostEqual(r["valor_incapacidad"], 3 * minimo / 30, places=2)
        self.assertAlmostEqual(r["total_a_pagar"], minimo, places=2)
        self.assertAlmostEqual(r["diferencia_vs_salario"], 0.0, places=2)
        self.assertIn("piso de 1 SMLMV", r["observaciones"])

    def test_incapacidad_que_cruza_tramo_baja_el_neto(self):
        """ANA arrastra 88 días de un episodio CONTINUO: de los 3 nuevos,
        2 van al 66,67% y 1 al 50%."""
        from app.services.nomina_report import armar_reporte
        r = armar_reporte(
            [self._fila(dias_incapacidad=3.0, valor_extras=31_250.0)],
            {"1": [self._incap("2026-05-01", "2026-07-27", 88.0, en_periodo=False),
                   self._incap("2026-07-28", "2026-07-30", 3.0)]},
            SMLMV,
        )[0]
        self.assertEqual(r["dias_efectivos"], 27)
        self.assertAlmostEqual(r["valor_incapacidad"], 2 * DIA_66 + 1 * DIA_50, places=2)
        self.assertAlmostEqual(r["total_a_pagar"], SALARIO / 30 * 27 + r["valor_incapacidad"] + 31_250, places=2)
        self.assertLess(r["diferencia_vs_salario"], 0)

    def test_los_dias_no_remunerados_descuentan(self):
        from app.services.nomina_report import armar_reporte
        r = armar_reporte([self._fila(dias_no_remunerados=2.0)], {}, SMLMV)[0]
        self.assertEqual(r["dias_efectivos"], 28)
        self.assertAlmostEqual(r["total_a_pagar"], SALARIO / 30 * 28, places=2)

    def test_vacaciones_no_reducen_el_neto(self):
        """No entran ni como incapacidad ni como no remunerados: el empleado
        sigue cobrando. Si esto se rompe, a quien toma vacaciones se le
        descontaria el sueldo."""
        from app.services.nomina_report import armar_reporte
        r = armar_reporte([self._fila()], {}, SMLMV)[0]
        self.assertAlmostEqual(r["diferencia_vs_salario"], 0.0, places=2)

    def test_sin_salario_se_avisa_y_la_diferencia_queda_vacia(self):
        from app.services.nomina_report import armar_reporte
        r = armar_reporte(
            [self._fila(salario=None, dias_incapacidad=2.0)],
            {"1": [self._incap("2026-08-01", "2026-08-02", 2.0)]}, SMLMV)[0]
        self.assertEqual(r["total_a_pagar"], 0.0)
        self.assertIsNone(r["diferencia_vs_salario"])
        self.assertIn("sin salario", r["observaciones"].lower())

    def test_novedades_por_encima_de_un_mes_no_dan_devengado_negativo(self):
        from app.services.nomina_report import armar_reporte
        r = armar_reporte(
            [self._fila(dias_incapacidad=40.0)],
            {"1": [self._incap("2026-08-01", "2026-09-09", 40.0)]}, SMLMV)[0]
        self.assertEqual(r["dias_efectivos"], 0)
        self.assertGreaterEqual(r["salario_devengado"], 0)
        self.assertIn("mas de un mes", r["observaciones"])


class TestLibroNomina(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
            import xlsxwriter  # noqa: F401
        except ImportError as exc:
            raise unittest.SkipTest(f"deps de Excel no disponibles ({exc})")

    def test_el_libro_lleva_el_aviso_de_alcance(self):
        """Quien abra el archivo debe leer que NO es una liquidación oficial
        antes que las cifras."""
        import openpyxl
        from app.services.nomina_report import armar_reporte
        from app.services.excel_report import construir_libro_nomina
        filas = armar_reporte(
            [dict(cedula="1", nombre_empleado="ANA", area="NOMINA", cargo="AUX",
                  salario=SALARIO, valor_extras=0.0, dias_incapacidad=0.0,
                  dias_no_remunerados=0.0, valor_otros_pagos=0.0)], {}, SMLMV)
        ws = openpyxl.load_workbook(construir_libro_nomina(filas, "2026-08")).active
        self.assertTrue((ws.cell(1, 1).value or "").startswith("Prenómina de apoyo"))
        self.assertIn("ARL", ws.cell(1, 1).value)
        self.assertEqual(ws.cell(ws.max_row, 1).value, "TOTAL")

    def test_el_libro_deja_constancia_del_piso_aplicado(self):
        """El SMLMV cambia cada año: sin dejarlo escrito no hay forma de
        auditar meses después con qué piso se liquidó esta prenómina."""
        import openpyxl
        from app.services.nomina_report import armar_reporte
        from app.services.excel_report import construir_libro_nomina
        filas = armar_reporte(
            [dict(cedula="1", nombre_empleado="ANA", area="NOMINA", cargo="AUX",
                  salario=SALARIO, valor_extras=0.0, dias_incapacidad=0.0,
                  dias_no_remunerados=0.0, valor_otros_pagos=0.0)], {}, SMLMV)
        libro = construir_libro_nomina(filas, "2024-08", smlmv=SMLMV, anio_smlmv=2024)
        ws = openpyxl.load_workbook(libro).active
        constancia = ws.cell(3, 1).value or ""
        self.assertIn("SMLMV 2024", constancia)
        self.assertIn("1,300,000", constancia)


class TestEpisodiosContinuos(unittest.TestCase):
    """Dos incapacidades separadas son enfermedades distintas: cada una vuelve
    a contar desde el día 1. Solo las continuas (prórrogas) acumulan."""

    def _reg(self, ini, fin, dias, en_periodo=True):
        return dict(fecha_inicio=ini, fecha_fin=fin, dias=dias, en_periodo=en_periodo)

    def test_contiguas_son_un_solo_episodio(self):
        from app.services.nomina_report import agrupar_episodios
        eps = agrupar_episodios([self._reg("2026-08-01", "2026-08-05", 5),
                                 self._reg("2026-08-06", "2026-08-08", 3)])
        self.assertEqual(len(eps), 1)

    def test_un_solo_dia_de_hueco_ya_separa(self):
        from app.services.nomina_report import agrupar_episodios
        eps = agrupar_episodios([self._reg("2026-08-01", "2026-08-05", 5),
                                 self._reg("2026-08-07", "2026-08-09", 3)])
        self.assertEqual(len(eps), 2)

    def test_solapadas_son_un_episodio(self):
        from app.services.nomina_report import agrupar_episodios
        eps = agrupar_episodios([self._reg("2026-08-01", "2026-08-10", 10),
                                 self._reg("2026-08-05", "2026-08-12", 8)])
        self.assertEqual(len(eps), 1)

    def test_separadas_no_llegan_al_tramo_del_50(self):
        """El caso que motivó el cambio: 100 días en dos episodios distintos
        se pagan todos al 66,67%, no al 50%."""
        from app.services.nomina_report import liquidar_incapacidades_empleado
        r = liquidar_incapacidades_empleado(
            [self._reg("2026-01-01", "2026-02-19", 50),
             self._reg("2026-06-01", "2026-07-20", 50)], SALARIO, SMLMV)
        self.assertEqual(r["dias_50"], 0)
        self.assertEqual(r["dias_66"], 100)
        self.assertAlmostEqual(r["valor"], 100 * DIA_66, places=2)

    def test_continuas_si_cruzan_al_50(self):
        from app.services.nomina_report import liquidar_incapacidades_empleado
        r = liquidar_incapacidades_empleado(
            [self._reg("2026-01-01", "2026-02-19", 50),
             self._reg("2026-02-20", "2026-04-10", 50)], SALARIO, SMLMV)
        self.assertEqual(r["dias_66"], 90)
        self.assertEqual(r["dias_50"], 10)
        self.assertAlmostEqual(r["valor"], 90 * DIA_66 + 10 * DIA_50, places=2)

    def test_lo_de_fuera_del_periodo_no_se_cobra_pero_acumula(self):
        from app.services.nomina_report import liquidar_incapacidades_empleado
        r = liquidar_incapacidades_empleado(
            [self._reg("2026-05-01", "2026-07-27", 88, en_periodo=False),
             self._reg("2026-07-28", "2026-07-30", 3)], SALARIO, SMLMV)
        self.assertEqual(r["dias_66"] + r["dias_50"], 3)   # solo se cobran 3
        self.assertEqual(r["dias_50"], 1)                   # pero arrastra el acumulado
